# -*- coding: utf-8 -*-
"""
当天订单梳理 - 自动处理Excel表格
功能：
1. 用户输入源文件和目标文件路径
2. 空值填充（指定列除外）
3. 日期格式转换
4. 按条件分离数据并追加到目标文件的不同工作表
5. 复制目标文件的格式
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font

import os

print("=" * 50)
print("    当天订单梳理 - 自动化处理工具")
print("=" * 50)
print()

# 用户输入源文件路径
print("请输入源文件路径 (当天订单.xls):")
print("示例: D:\\共享\\当天订单.xls")
SOURCE_FILE = input("路径: ").strip()
if not SOURCE_FILE:
    SOURCE_FILE = r"D:\共享\当天订单.xls"
    print(f"使用默认路径: {SOURCE_FILE}")

# 用户输入目标文件路径
print()
print("请输入目标文件路径 (系统销售订单汇总.xlsx):")
print("示例: D:\\共享\\02-系统销售订单汇总-2026.03.31（系统版）.xlsx")
TARGET_FILE = input("路径: ").strip()
if not TARGET_FILE:
    TARGET_FILE = r"D:\共享\02-系统销售订单汇总-2026.03.31（系统版）.xlsx"
    print(f"使用默认路径: {TARGET_FILE}")

# 检查文件是否存在
if not os.path.exists(SOURCE_FILE):
    print(f"\n错误: 源文件不存在: {SOURCE_FILE}")
    input("按回车键退出...")
    exit(1)

if not os.path.exists(TARGET_FILE):
    print(f"\n错误: 目标文件不存在: {TARGET_FILE}")
    input("按回车键退出...")
    exit(1)

print()
print("-" * 50)

# 配置参数
SKIP_FILL_COLUMNS = ['客户订单号']  # 不填充空值的列
BK_KEYWORD = 'BK'  # 购货单位包含此关键词的数据追加到"其他销售订单"

# 1. 读取源文件
print(f"\n[1] 读取源文件: {SOURCE_FILE}")
df = pd.read_excel(SOURCE_FILE)
print(f"    源文件总行数: {len(df)}")

# 加载目标文件以获取列名
print(f"\n    加载目标文件获取列名: {TARGET_FILE}")
wb = load_workbook(TARGET_FILE)
ws_target = wb['其他销售订单']  # 随便取一个工作表，列名应该是一样的
target_cols = [ws_target.cell(1, col).value for col in range(1, 14)]

# 将源文件的列名(前13列)改为目标文件的列名
if len(df.columns) >= 13:
    print(f"    将源文件前13列的列名修改为目标文件的列名...")
    df.columns = target_cols + list(df.columns[13:])
else:
    print(f"    ⚠️ 源文件列数小于13，可能无法完全匹配！")

col_date = target_cols[0]      # 订单日期 (A列)
col_ship = target_cols[1]      # 发货方式 (B列)
col_customer = target_cols[2]  # 购货单位 (C列)
col_dept = target_cols[3]      # 部门 (D列)
col_sales = target_cols[4]     # 业务员 (E列)
col_product = target_cols[5]   # 产品名称 (F列)
col_spec = target_cols[6]      # 规格型号 (G列)
col_qty = target_cols[7]       # 数量 (H列)
col_unit = target_cols[8]      # 单位 (I列)
col_price = target_cols[9]     # 含税单价 (J列)
col_amount = target_cols[10]   # 总金额 (K列)
col_summary = target_cols[11]  # 摘要 (L列)
col_orderid = target_cols[12]  # 客户订单号 (M列)

# 配置参数更新，使用目标文件M列的名字
SKIP_FILL_COLUMNS = [str(col_orderid)]  # 不填充空值的列

# 2. 空值填充
print("\n[2] 执行空值填充...")

for col in df.columns:
    if str(col).strip() not in [c.strip() for c in SKIP_FILL_COLUMNS]:
        df[col] = df[col].ffill().bfill()
print(f"    空值填充完成 (M列'{col_orderid}'不填充)")

# 3. 日期格式：保持为日期值，设置为 YYYY/M/D 显示格式
print("\n[3] 处理日期格式（保持日期值，设置 YYYY/M/D 显示格式）...")
df[col_date] = pd.to_datetime(df[col_date], errors='coerce')
print("    日期格式处理完成（已转为日期值，支持筛选树分组）")

# 4. 分离BK和非BK数据
print("\n[4] 分离数据...")
df_bk = df[df[col_customer].astype(str).str.contains(BK_KEYWORD, na=False)].copy()
df_other = df[~df[col_customer].astype(str).str.contains(BK_KEYWORD, na=False)].copy()
print(f"    BK数据 (追加到'其他销售订单'): {len(df_bk)}行")
print(f"    其他数据 (追加到'开票销售订单'): {len(df_other)}行")

# 5. 加载目标文件（已在第1步提前加载）
print(f"\n[5] 准备写入目标文件: {TARGET_FILE}")

# 6. 处理"其他销售订单" - 追加BK数据
print("\n[6] 处理'其他销售订单'工作表...")
ws_other = wb['其他销售订单']
last_row_other = 0
for row in range(ws_other.max_row, 0, -1):
    if any(ws_other.cell(row, col).value is not None for col in range(1, 14)):
        last_row_other = row
        break
print(f"    最后数据行: {last_row_other}")

src_row_other = last_row_other if last_row_other > 1 else 1

for idx, row_data in df_bk.iterrows():
    new_row = last_row_other + 1 + df_bk.index.get_loc(idx)
    # 日期列：保持日期值，设置为 YYYY/M/D 显示格式
    date_cell = ws_other.cell(new_row, 1)
    date_cell.value = row_data[col_date]
    date_cell.number_format = 'YYYY/M/D'
    ws_other.cell(new_row, 2).value = row_data[col_ship]
    ws_other.cell(new_row, 3).value = row_data[col_customer]
    ws_other.cell(new_row, 4).value = row_data[col_dept]
    ws_other.cell(new_row, 5).value = row_data[col_sales]
    ws_other.cell(new_row, 6).value = row_data[col_product]
    ws_other.cell(new_row, 7).value = row_data[col_spec]
    ws_other.cell(new_row, 8).value = row_data[col_qty]
    ws_other.cell(new_row, 9).value = row_data[col_unit]
    ws_other.cell(new_row, 10).value = row_data[col_price]
    ws_other.cell(new_row, 11).value = row_data[col_amount]
    ws_other.cell(new_row, 12).value = row_data[col_summary]
    ws_other.cell(new_row, 13).value = row_data[col_orderid]

    for col in range(1, 14):
        src_cell = ws_other.cell(src_row_other, col)
        dst_cell = ws_other.cell(new_row, col)
        dst_cell.font = Font(name='宋体', size=src_cell.font.size if src_cell.font else 11)
        if src_cell.alignment:
            dst_cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical)

    # 统一行高为24
    ws_other.row_dimensions[new_row].height = 24

print(f"    已追加 {len(df_bk)} 行")

# 统一其他销售订单所有行高为24
for row in range(1, ws_other.max_row + 1):
    ws_other.row_dimensions[row].height = 24
print(f"    已将所有行高设为24")

# 7. 处理"开票销售订单" - 追加非BK数据
print("\n[7] 处理'开票销售订单'工作表...")
ws_kp = wb['开票销售订单']
last_row_kp = 0
for row in range(ws_kp.max_row, 0, -1):
    if any(ws_kp.cell(row, col).value is not None for col in range(1, 14)):
        last_row_kp = row
        break
print(f"    最后数据行: {last_row_kp}")

src_row_kp = last_row_kp if last_row_kp > 1 else 1
sample_cell = ws_kp.cell(src_row_kp, 1)
border_style = sample_cell.border.left.style if sample_cell.border and sample_cell.border.left else 'thin'

for idx, row_data in df_other.iterrows():
    new_row = last_row_kp + 1 + df_other.index.get_loc(idx)
    # 日期列：保持日期值，设置为 YYYY/M/D 显示格式
    date_cell = ws_kp.cell(new_row, 1)
    date_cell.value = row_data[col_date]
    date_cell.number_format = 'YYYY/M/D'
    ws_kp.cell(new_row, 2).value = row_data[col_ship]
    ws_kp.cell(new_row, 3).value = row_data[col_customer]
    ws_kp.cell(new_row, 4).value = row_data[col_dept]
    ws_kp.cell(new_row, 5).value = row_data[col_sales]
    ws_kp.cell(new_row, 6).value = row_data[col_product]
    ws_kp.cell(new_row, 7).value = row_data[col_spec]
    ws_kp.cell(new_row, 8).value = row_data[col_qty]
    ws_kp.cell(new_row, 9).value = row_data[col_unit]
    ws_kp.cell(new_row, 10).value = row_data[col_price]
    ws_kp.cell(new_row, 11).value = f'=H{new_row}*J{new_row}'
    ws_kp.cell(new_row, 12).value = row_data[col_summary]
    ws_kp.cell(new_row, 13).value = row_data[col_orderid]

    for col in range(1, 14):
        src_cell = ws_kp.cell(src_row_kp, col)
        dst_cell = ws_kp.cell(new_row, col)
        side = Side(style=border_style, color='000000')
        dst_cell.border = Border(left=side, right=side, top=side, bottom=side)
        dst_cell.font = Font(name='宋体', size=src_cell.font.size if src_cell.font else 11)
        if src_cell.alignment:
            dst_cell.alignment = Alignment(horizontal=src_cell.alignment.horizontal, vertical=src_cell.alignment.vertical)

    # 统一行高为24
    ws_kp.row_dimensions[new_row].height = 24

print(f"    已追加 {len(df_other)} 行")

# 统一开票销售订单所有行高为24
for row in range(1, ws_kp.max_row + 1):
    ws_kp.row_dimensions[row].height = 24
print(f"    已将所有行高设为24")

# 8. 保存
print("\n[8] 保存文件...")
wb.save(TARGET_FILE)

print()
print("=" * 50)
print("    处理完成!")
print("=" * 50)
print(f"  - 其他销售订单: {len(df_bk)}行 (从第{last_row_other+1}行开始)")
print(f"  - 开票销售订单: {len(df_other)}行 (从第{last_row_kp+1}行开始)")
print()
input("按回车键退出...")
